import streamlit as st
from datetime import date, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

st.set_page_config(page_title="Cash Advance Form Builder", page_icon="💵", layout="wide")


# ---------- Helpers ----------
def fmt_date(value):
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if isinstance(value, date):
        return value.strftime("%Y/%m/%d")
    return str(value)


def fmt_currency(value):
    try:
        if value is None or value == "":
            return ""
        return f"PHP {float(value):,.2f}"
    except Exception:
        return str(value)


def safe_text(value):
    return "" if value is None else str(value)


def build_workbook(data: dict) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Advance Form"

    # ---------- Theme ----------
    dark = "1F4E78"
    mid = "D9E7F5"
    light = "F7FAFC"
    note_fill = "FFF4CC"
    white = "FFFFFF"
    black = "222222"

    thin_gray = Side(style="thin", color="9AA4B2")
    medium_dark = Side(style="medium", color="4B5563")

    title_font = Font(name="Calibri", size=16, bold=True, color=black)
    section_font = Font(name="Calibri", size=11, bold=True, color=white)
    label_font = Font(name="Calibri", size=10, bold=True, color=black)
    value_font = Font(name="Calibri", size=10, color=black)
    note_font = Font(name="Calibri", size=9, italic=True, color=black)
    sign_font = Font(name="Calibri", size=9, bold=True, color=black)
    mini_font = Font(name="Calibri", size=8, color="666666")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    top_left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def set_range_border(cell_range, border):
        rows = ws[cell_range]
        for row in rows:
            for cell in row:
                cell.border = border

    def merge_value(label_row, label_col, label, value, value_end_col, height=22):
        ws.cell(label_row, label_col, label)
        ws.cell(label_row, label_col).font = label_font
        ws.cell(label_row, label_col).alignment = left
        ws.cell(label_row, label_col).fill = PatternFill("solid", fgColor=light)
        ws.merge_cells(start_row=label_row, start_column=label_col + 1, end_row=label_row, end_column=value_end_col)
        vcell = ws.cell(label_row, label_col + 1, safe_text(value))
        vcell.font = value_font
        vcell.alignment = left
        ws.row_dimensions[label_row].height = height
        for col in range(label_col, value_end_col + 1):
            ws.cell(label_row, col).border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # ---------- Layout ----------
    widths = {
        "A": 14, "B": 14, "C": 14, "D": 14,
        "E": 14, "F": 14, "G": 14, "H": 14,
        "I": 14, "J": 14, "K": 14, "L": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A1"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.45, bottom=0.45, header=0.2, footer=0.2)
    ws.print_title_rows = "$1:$3"
    ws.print_area = "A1:L42"

    # ---------- Title ----------
    ws.merge_cells("A1:J2")
    ws["A1"] = "Cash Advance Form"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("K1:L2")
    ws["K1"] = f"Control No.\n{safe_text(data['control_no'])}"
    ws["K1"].font = label_font
    ws["K1"].alignment = center_wrap
    ws["K1"].fill = PatternFill("solid", fgColor=mid)
    set_range_border("K1:L2", Border(left=medium_dark, right=medium_dark, top=medium_dark, bottom=medium_dark))
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22

    def section_header(row, title):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        c = ws.cell(row, 1, title)
        c.font = section_font
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = left
        set_range_border(f"A{row}:L{row}", Border(left=medium_dark, right=medium_dark, top=medium_dark, bottom=medium_dark))
        ws.row_dimensions[row].height = 22

    def signature_block(start_row, role1, role2, role3):
        labels = ["Prepared by", "Confirmed by", "Approved by"]
        roles = [role1, role2, role3]
        starts = [1, 5, 9]
        for idx, start_col in enumerate(starts):
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
            ws.cell(start_row, start_col, labels[idx])
            ws.cell(start_row, start_col).font = mini_font
            ws.cell(start_row, start_col).alignment = left

            ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=start_row + 2, end_column=start_col + 3)
            cell = ws.cell(start_row + 1, start_col, roles[idx])
            cell.font = sign_font
            cell.alignment = center
            rng = f"{get_column_letter(start_col)}{start_row+1}:{get_column_letter(start_col+3)}{start_row+2}"
            set_range_border(rng, Border(top=medium_dark))
        ws.row_dimensions[start_row].height = 18
        ws.row_dimensions[start_row + 1].height = 20
        ws.row_dimensions[start_row + 2].height = 18

    # ---------- Section 1 ----------
    section_header(4, "1. Application")
    merge_value(5, 1, "Applicant", data["applicant"], 4)
    merge_value(5, 5, "Department", data["department"], 8)
    merge_value(5, 9, "Office", data["office"], 12)

    merge_value(6, 1, "Application Date", fmt_date(data["application_date"]), 4)
    merge_value(6, 5, "Borrow Date", fmt_date(data["borrow_date"]), 8)
    merge_value(6, 9, "Planned Settlement Date", fmt_date(data["planned_settlement_date"]), 12)

    merge_value(7, 1, "Application Amount", fmt_currency(data["application_amount"]), 12)

    ws.cell(8, 1, "Purpose / Reason")
    ws.cell(8, 1).font = label_font
    ws.cell(8, 1).alignment = left
    ws.cell(8, 1).fill = PatternFill("solid", fgColor=light)
    ws.merge_cells("B8:L10")
    ws["B8"] = safe_text(data["application_reason"])
    ws["B8"].font = value_font
    ws["B8"].alignment = top_left_wrap
    set_range_border("A8:L10", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[9].height = 24
    ws.row_dimensions[10].height = 24

    ws.merge_cells("A11:L11")
    ws["A11"] = "Note: If settlement is delayed beyond the planned date, the outstanding amount may be deducted from the next payroll."
    ws["A11"].font = note_font
    ws["A11"].alignment = left
    ws["A11"].fill = PatternFill("solid", fgColor=note_fill)
    set_range_border("A11:L11", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))
    signature_block(12, "Recipient", "Immediate Supervisor", "Cashier")

    # ---------- Section 2 ----------
    section_header(16, "2. Settlement")
    merge_value(17, 1, "Settlement Date", fmt_date(data["settlement_date"]), 12)
    merge_value(18, 1, "Cash Advance Amount", fmt_currency(data["cash_advance_amount"]), 4)
    merge_value(18, 5, "Total Receipts", fmt_currency(data["receipt_total"]), 8)
    merge_value(18, 9, "Change", fmt_currency(data["change_amount"]), 12)
    merge_value(19, 1, "Unsettled Amount", fmt_currency(data["unsettled_amount"]), 12)

    ws.merge_cells("A20:L20")
    ws["A20"] = "Note: This is to certify that the above settlement is true and correct."
    ws["A20"].font = note_font
    ws["A20"].alignment = left
    ws["A20"].fill = PatternFill("solid", fgColor=note_fill)
    set_range_border("A20:L20", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))
    signature_block(21, "Recipient", "Cashier", "CFO")

    # ---------- Section 3 ----------
    section_header(25, "3. Repayment")
    merge_value(26, 1, "Payer", data["repayer"], 4)
    merge_value(26, 5, "Department", data["repayer_department"], 8)
    merge_value(26, 9, "Office", data["repayer_office"], 10)
    ws.cell(26, 11, "Entry Date")
    ws.cell(26, 11).font = label_font
    ws.cell(26, 11).fill = PatternFill("solid", fgColor=light)
    ws.cell(26, 11).alignment = left
    ws.cell(26, 12, fmt_date(data["repayment_entry_date"]))
    ws.cell(26, 12).font = value_font
    ws.cell(26, 12).alignment = left
    set_range_border("A26:L26", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))

    merge_value(27, 1, "Total Repayment Amount", fmt_currency(data["repayment_total"]), 12)
    merge_value(28, 1, "Repayment Method", data["repayment_method"], 12)
    merge_value(29, 1, "Number of Installments", data["installments"], 12)
    merge_value(30, 1, "First Deduction Date", fmt_date(data["first_deduction_date"]), 12)

    ws.merge_cells("A31:L31")
    ws["A31"] = "Note: I hereby agree to repay the above amount according to the stated terms."
    ws["A31"].font = note_font
    ws["A31"].alignment = left
    ws["A31"].fill = PatternFill("solid", fgColor=note_fill)
    set_range_border("A31:L31", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))
    signature_block(32, "Recipient", "Cashier", "CFO")

    # Footer memo area
    ws.merge_cells("A36:L37")
    ws["A36"] = "Remarks"
    ws["A36"].font = label_font
    ws["A36"].alignment = top_left_wrap
    ws["A36"].fill = PatternFill("solid", fgColor=light)
    set_range_border("A36:L37", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))

    # Outer frame
    set_range_border("A4:L34", Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray))

    # Keep everything readable as one sheet
    for row in range(38, 43):
        ws.row_dimensions[row].height = 6

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def default_state():
    today = date.today()
    return {
        "control_no": f"CA-{today.strftime('%Y%m%d')}-001",
        "applicant": "",
        "department": "",
        "office": "",
        "application_date": today,
        "borrow_date": today,
        "planned_settlement_date": today,
        "application_amount": 0,
        "application_reason": "",
        "settlement_date": today,
        "cash_advance_amount": 0,
        "receipt_total": 0,
        "change_amount": 0,
        "repayer": "",
        "repayer_department": "",
        "repayer_office": "",
        "repayment_entry_date": today,
        "repayment_total": 0,
        "repayment_method": "Payroll Deduction",
        "installments": 1,
        "first_deduction_date": today,
    }


if "form_data" not in st.session_state:
    st.session_state.form_data = default_state()


def set_sample_data():
    today = date.today()
    st.session_state.form_data = {
        "control_no": f"CA-{today.strftime('%Y%m%d')}-001",
        "applicant": "Juan Dela Cruz",
        "department": "Sales Department",
        "office": "Davao",
        "application_date": today,
        "borrow_date": today,
        "planned_settlement_date": today,
        "application_amount": 30000,
        "application_reason": "Cash advance request for travel expenses, accommodation, and other reimbursable business expenses.",
        "settlement_date": today,
        "cash_advance_amount": 30000,
        "receipt_total": 24500,
        "change_amount": 3000,
        "repayer": "Juan Dela Cruz",
        "repayer_department": "Sales Department",
        "repayer_office": "Davao",
        "repayment_entry_date": today,
        "repayment_total": 2500,
        "repayment_method": "Payroll Deduction",
        "installments": 1,
        "first_deduction_date": today,
    }


# ---------- UI ----------
st.title("💵 Cash Advance Form Builder")
st.caption("Enter the application, settlement, and repayment details on one screen, then export them to a single-sheet Excel file.")

col_a, col_b = st.columns([3, 1])
with col_b:
    if st.button("Load Sample Data", use_container_width=True):
        set_sample_data()
        st.rerun()

with st.form("cash_advance_form"):
    st.subheader("Basic Information")
    c1, c2 = st.columns(2)
    with c1:
        control_no = st.text_input("Control No.", value=st.session_state.form_data["control_no"])
    with c2:
        applicant = st.text_input("Applicant", value=st.session_state.form_data["applicant"])

    c3, c4 = st.columns(2)
    with c3:
        department = st.text_input("Department", value=st.session_state.form_data["department"])
    with c4:
        office = st.text_input("Office", value=st.session_state.form_data["office"])

    st.subheader("1. Application")
    d1, d2, d3 = st.columns(3)
    with d1:
        application_date = st.date_input("Application Date", value=st.session_state.form_data["application_date"])
    with d2:
        borrow_date = st.date_input("Borrow Date", value=st.session_state.form_data["borrow_date"])
    with d3:
        planned_settlement_date = st.date_input("Planned Settlement Date", value=st.session_state.form_data["planned_settlement_date"])

    application_amount = st.number_input(
        "Application Amount (PHP)",
        min_value=0,
        step=1000,
        value=int(st.session_state.form_data["application_amount"]),
    )
    application_reason = st.text_area(
        "Purpose / Reason",
        value=st.session_state.form_data["application_reason"],
        height=100,
        placeholder="Example: travel expenses, item purchase, site support expenses, or other reimbursable business costs.",
    )

    st.subheader("2. Settlement")
    settlement_date = st.date_input("Settlement Date", value=st.session_state.form_data["settlement_date"])
    e1, e2, e3 = st.columns(3)
    with e1:
        cash_advance_amount = st.number_input(
            "Cash Advance Amount (PHP)", min_value=0, step=1000, value=int(st.session_state.form_data["cash_advance_amount"])
        )
    with e2:
        receipt_total = st.number_input(
            "Total Receipts (PHP)", min_value=0, step=1000, value=int(st.session_state.form_data["receipt_total"])
        )
    with e3:
        change_amount = st.number_input(
            "Change (PHP)", min_value=0, step=1000, value=int(st.session_state.form_data["change_amount"])
        )

    unsettled_amount = cash_advance_amount - receipt_total - change_amount
    if unsettled_amount > 0:
        st.warning(f"Unsettled Amount: {fmt_currency(unsettled_amount)}")
    elif unsettled_amount < 0:
        st.info(f"Additional settlement is required: {fmt_currency(abs(unsettled_amount))}")
    else:
        st.success("There is no unsettled amount.")

    st.subheader("3. Repayment")
    f1, f2, f3 = st.columns(3)
    with f1:
        repayer = st.text_input("Payer", value=st.session_state.form_data["repayer"])
    with f2:
        repayer_department = st.text_input("Payer Department", value=st.session_state.form_data["repayer_department"])
    with f3:
        repayer_office = st.text_input("Payer Office", value=st.session_state.form_data["repayer_office"])

    g1, g2, g3, g4 = st.columns(4)
    with g1:
        repayment_entry_date = st.date_input("Repayment Entry Date", value=st.session_state.form_data["repayment_entry_date"])
    with g2:
        repayment_total = st.number_input(
            "Total Repayment Amount (PHP)", min_value=0, step=1000, value=int(st.session_state.form_data["repayment_total"])
        )
    with g3:
        repayment_method = st.selectbox(
            "Repayment Method",
            ["COOP", "Payroll Deduction"],
            index=["COOP", "Payroll Deduction"].index(st.session_state.form_data["repayment_method"]),
        )
    with g4:
        installments = st.number_input(
            "Number of Installments", min_value=1, step=1, value=int(st.session_state.form_data["installments"])
        )

    first_deduction_date = st.date_input(
        "First Deduction Date",
        value=st.session_state.form_data["first_deduction_date"],
    )

    submitted = st.form_submit_button("Update Form Data", use_container_width=True)

if submitted:
    st.session_state.form_data = {
        "control_no": control_no,
        "applicant": applicant,
        "department": department,
        "office": office,
        "application_date": application_date,
        "borrow_date": borrow_date,
        "planned_settlement_date": planned_settlement_date,
        "application_amount": application_amount,
        "application_reason": application_reason,
        "settlement_date": settlement_date,
        "cash_advance_amount": cash_advance_amount,
        "receipt_total": receipt_total,
        "change_amount": change_amount,
        "unsettled_amount": unsettled_amount,
        "repayer": repayer,
        "repayer_department": repayer_department,
        "repayer_office": repayer_office,
        "repayment_entry_date": repayment_entry_date,
        "repayment_total": repayment_total,
        "repayment_method": repayment_method,
        "installments": installments,
        "first_deduction_date": first_deduction_date,
    }
    st.success("The form data has been updated.")

# Ensure computed field exists before preview/download
preview_data = st.session_state.form_data.copy()
preview_data["unsettled_amount"] = (
    preview_data.get("cash_advance_amount", 0)
    - preview_data.get("receipt_total", 0)
    - preview_data.get("change_amount", 0)
)

st.divider()
st.subheader("Preview")

p1, p2 = st.columns(2)
with p1:
    st.markdown("#### Application")
    st.write(f"**Control No.**: {preview_data['control_no']}")
    st.write(f"**Applicant**: {preview_data['applicant']}")
    st.write(f"**Department / Office**: {preview_data['department']} / {preview_data['office']}")
    st.write(f"**Application Date**: {fmt_date(preview_data['application_date'])}")
    st.write(f"**Borrow Date**: {fmt_date(preview_data['borrow_date'])}")
    st.write(f"**Planned Settlement Date**: {fmt_date(preview_data['planned_settlement_date'])}")
    st.write(f"**Application Amount**: {fmt_currency(preview_data['application_amount'])}")
    st.write(f"**Purpose / Reason**: {preview_data['application_reason']}" if preview_data['application_reason'] else "**Purpose / Reason**: -")

with p2:
    st.markdown("#### Settlement / Repayment")
    st.write(f"**Settlement Date**: {fmt_date(preview_data['settlement_date'])}")
    st.write(f"**Cash Advance Amount**: {fmt_currency(preview_data['cash_advance_amount'])}")
    st.write(f"**Total Receipts**: {fmt_currency(preview_data['receipt_total'])}")
    st.write(f"**Change**: {fmt_currency(preview_data['change_amount'])}")
    st.write(f"**Unsettled Amount**: {fmt_currency(preview_data['unsettled_amount'])}")
    st.write(f"**Payer**: {preview_data['repayer']}")
    st.write(f"**Repayment Method**: {preview_data['repayment_method']}")
    st.write(f"**Total Repayment Amount**: {fmt_currency(preview_data['repayment_total'])}")
    st.write(f"**Number of Installments**: {preview_data['installments']}")
    st.write(f"**First Deduction Date**: {fmt_date(preview_data['first_deduction_date'])}")

excel_bytes = build_workbook(preview_data)
st.download_button(
    label="📥 Download Excel (Single-Sheet Version)",
    data=excel_bytes,
    file_name=f"cash_advance_{preview_data['control_no'] or 'form'}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

st.info("The output file is a single-sheet layout designed with A4 printing in mind.")
